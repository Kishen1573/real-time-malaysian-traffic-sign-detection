import os
import csv
from datetime import datetime
import pytz

# Excel support
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ---------- Columns we write (UTC removed) ----------
CSV_HEADER = [
    "timestamp_myt", "date_myt", "frame_idx", "label", "confidence",
    "image_relpath",
    "blur_score", "contrast_score", "glare_ratio", "visibility_score",
    "low_visibility",
    "problem_hint", "jkr_action", "is_trusted",
]

LOW_VIS_COL_IDX = CSV_HEADER.index("low_visibility") + 1  # 1-based for Excel
HEADER_FILL = PatternFill("solid", fgColor="FFF3CD")
MYT = pytz.timezone("Asia/Kuala_Lumpur")


def _today_dir(base_dir: str) -> str:
    """Return today's detections directory (MYT) and ensure it exists."""
    date_str = datetime.now(MYT).strftime("%Y-%m-%d")
    day_dir = os.path.join(base_dir, date_str)
    os.makedirs(day_dir, exist_ok=True)
    return day_dir


def open_csv_for_today(base_dir: str) -> str:
    """Create the CSV for today if needed and return its path."""
    day_dir = _today_dir(base_dir)
    csv_path = os.path.join(day_dir, "detections.csv")
    if not os.path.exists(csv_path):
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADER)
    return csv_path


def open_xlsx_for_today(base_dir: str) -> str:
    """Create the XLSX for today if needed and return its path."""
    day_dir = _today_dir(base_dir)
    xlsx_path = os.path.join(day_dir, "detections.xlsx")
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "detections"
        ws.append(CSV_HEADER)
        # Style header row
        for col in range(1, len(CSV_HEADER) + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    return xlsx_path


def _row_from_dict(row_dict: dict):
    """Project a dict into our fixed column order."""
    return [row_dict.get(k, "") for k in CSV_HEADER]


def append_csv_row(csv_path: str, row_dict: dict):
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        csv.writer(f).writerow(_row_from_dict(row_dict))


def append_xlsx_row(xlsx_path: str, row_dict: dict):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.append(_row_from_dict(row_dict))
    r = ws.max_row  # the row we just appended

    # If low_visibility is TRUE, color the entire row light red
    val = str(ws.cell(row=r, column=LOW_VIS_COL_IDX).value).strip().lower()
    if val in ("true", "1", "yes"):
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        for c in range(1, len(CSV_HEADER) + 1):
            ws.cell(row=r, column=c).fill = red_fill

    wb.save(xlsx_path)
